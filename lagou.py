# !/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time     : 2018/8/14 20:30
# @Author   : zhe
# @FileName : lagou.py
# @Project  : PyCharm


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from collections import OrderedDict
from random import randint
import os
import time
import openpyxl

BASE_DIR = os.path.dirname(__file__)

chromedriver_path = os.path.join(BASE_DIR, 'chromedriver.exe')


class Lagou():

    def __init__(self, driver_path, start_url, page_max, xlsx_name,
                 sheet_name):
        self.driver_path = driver_path
        self.start_url = start_url
        self.row = 1
        self.xlsx_name = xlsx_name + '.xlsx'
        self.sheet_name = sheet_name
        self.page_max = page_max

    def job_info(self):
        info = OrderedDict()
        jobs = []
        jobs.append(self.driver.find_element_by_xpath(
            '//li[@class="con_list_item first_row default_list"]'))
        jobs.extend(self.driver.find_elements_by_xpath(
            '//li[@class="con_list_item default_list"]'))
        for job in jobs:
            index = job.get_attribute('data-index')
            info[index] = OrderedDict()
            info[index]['job'] = job.find_element_by_xpath(
                './/a[@class="position_link"]/h3').text
            info[index]['addr'] = job.find_element_by_xpath(
                './/a[@class="position_link"]/span/em').text
            info[index]['company_name'] = job.find_element_by_xpath(
                './/div[@class="company_name"]/a').text
            info[index]['money'] = job.find_element_by_xpath(
                './/span[@class="money"]').text
            info[index]['experience'] = job.find_element_by_xpath(
                './/div[@class="p_bot"]/div[@class="li_b_l"]').text
            info[index]['industry'] = job.find_element_by_xpath(
                './/div[@class="industry"]').text
            info[index]['cats'] = \
                ','.join([cat.text for cat in job.find_elements_by_xpath(
                    './/div[@class="list_item_bot"]/div[@class="li_b_l"]/span')])
            info[index]['benefits'] = job.find_element_by_xpath(
                './/div[@class="list_item_bot"]/div[@class="li_b_r"]').text
        return info

    def xlsx_init(self):
        if os.path.exists(os.path.join(BASE_DIR, self.xlsx_name)):
            self.xlsx = openpyxl.load_workbook(
                os.path.join(BASE_DIR, self.xlsx_name))
        else:
            self.xlsx = openpyxl.Workbook()
            del self.xlsx['Sheet']
        try:
            sheet = self.xlsx[self.sheet_name]
            self.xlsx.remove(sheet)
        except:
            pass
        finally:
            self.xlsx.create_sheet(self.sheet_name)
            self.sheet = self.xlsx[self.sheet_name]

    def write_to_xlsx(self, info):
        for k1, v1 in info.items():
            column = 1
            for k2, v2 in v1.items():
                self.sheet.cell(row=self.row, column=column).value = v2
                column += 1
            self.row += 1

    def xlsx_save(self):
        self.xlsx.save(os.path.join(BASE_DIR, self.xlsx_name))

    def main(self):
        self.driver = webdriver.Chrome(executable_path=self.driver_path)
        self.driver.get(self.start_url)
        self.xlsx_init()
        element = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.XPATH,
                                            '//li[@class="con_list_item first_row default_list"]')))
        self.driver.execute_script(
            "window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(randint(1, 3))
        info = self.job_info()
        self.write_to_xlsx(info)
        page_list = self.driver.find_elements_by_xpath(
            '//div[@class="pager_container"]/span[@class="pager_not_current"]')
        page_max = 0
        for page in page_list:
            num = int(page.get_attribute('page'))
            if num > page_max:
                page_max = num
        if page_max > self.page_max:
            page_max = self.page_max
        try:
            for i in range(2, page_max + 1):
                self.driver.find_element_by_xpath(
                    '//div[@class="pager_container"]/span[@class="pager_not_current" and @page="%s"]' % str(
                        i)
                ).click()
                elements = WebDriverWait(self.driver, 10).until(
                    EC.presence_of_all_elements_located((By.XPATH,
                                                         '//li[@class="con_list_item default_list"]')))
                self.driver.execute_script(
                    "window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(randint(1, 3))
                info = self.job_info()
                self.write_to_xlsx(info)
        except Exception as e:
            print(e)
        finally:
            self.driver.close()
            self.xlsx_save()


if __name__ == '__main__':

    # 成都前端和Python职位爬取
    html_1_3 = Lagou(
        chromedriver_path,
        '''https://www.lagou.com/jobs/list_%E5%89%8D%E7%AB%AF
        ?gj=3%E5%B9%B4%E5%8F%8A%E4%BB%A5%E4%B8%8B&px=new
        &city=%E6%88%90%E9%83%BD#order''',
        10,
        'html',
        '1-3'
    )
    html_all = Lagou(
        chromedriver_path,
        '''https://www.lagou.com/jobs/list_%E5%89%8D%E7%AB%AF
        ?px=new&city=%E6%88%90%E9%83%BD#order''',
        10,
        'html',
        'all'
    )
    python_all = Lagou(
        chromedriver_path,
        '''https://www.lagou.com/jobs/list_python
        ?px=new&city=%E6%88%90%E9%83%BD#order''',
        10,
        'python',
        'all'
    )

    python_1_3 = Lagou(
        chromedriver_path,
        '''https://www.lagou.com/jobs/list_python
        ?px=new&gj=3%E5%B9%B4%E5%8F%8A%E4%BB%A5%E4%B8%8B
        &city=%E6%88%90%E9%83%BD#filterBox''',
        10,
        'python',
        '1-3'
    )

    tasks = [python_all, python_1_3, html_all, html_1_3]

    for task in tasks:
        task.main()
